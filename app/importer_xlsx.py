from __future__ import annotations

from io import BytesIO
from typing import Any

import openpyxl

REQ_TOPICS = ["topic_id", "section", "title_ru", "title_en", "order", "status", "updated_at", "notes"]
REQ_CARDS = ["card_id", "topic_id", "concept", "brief", "example", "when_use", "pitfalls", "keywords", "status", "updated_at", "search_text"]

def _sheet_to_dicts(ws) -> list[dict[str, Any]]:
  rows = list(ws.iter_rows(values_only=True))
  if not rows:
    return []
  header = [str(c).strip() if c is not None else "" for c in rows[0]]
  out = []
  for r in rows[1:]:
    if all(v is None or str(v).strip() == "" for v in r):
      continue
    obj = {}
    for i, key in enumerate(header):
      if not key:
        continue
      obj[key] = r[i] if i < len(r) else None
    out.append(obj)
  return out

def _require_cols(objs: list[dict], required: list[str], sheet: str) -> None:
  if not objs:
    raise ValueError(f"Лист '{sheet}' пустой")
  cols = set(objs[0].keys())
  missing = [c for c in required if c not in cols]
  if missing:
    raise ValueError(f"В листе '{sheet}' не хватает колонок: {missing}")

def _unique_or_fail(objs: list[dict], key: str, sheet: str) -> None:
  seen = set()
  dups = set()
  for o in objs:
    v = str(o.get(key, "")).strip()
    if not v:
      continue
    if v in seen:
      dups.add(v)
    seen.add(v)
  if dups:
    raise ValueError(f"Дубликаты {key} в листе '{sheet}': {sorted(list(dups))[:10]} ...")

def parse_xlsx_bytes(xlsx_bytes: bytes) -> dict:
  wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)

  if "topics" not in wb.sheetnames or "cards" not in wb.sheetnames:
    raise ValueError("Ожидаются вкладки 'topics' и 'cards'")

  topics_raw = _sheet_to_dicts(wb["topics"])
  cards_raw = _sheet_to_dicts(wb["cards"])

  _require_cols(topics_raw, REQ_TOPICS, "topics")
  _require_cols(cards_raw, REQ_CARDS, "cards")

  _unique_or_fail(topics_raw, "topic_id", "topics")
  _unique_or_fail(cards_raw, "card_id", "cards")

  # tags/card_tags/aliases/links — опционально
  tags_raw = _sheet_to_dicts(wb["tags"]) if "tags" in wb.sheetnames else []
  card_tags_raw = _sheet_to_dicts(wb["card_tags"]) if "card_tags" in wb.sheetnames else []
  aliases_raw = _sheet_to_dicts(wb["aliases"]) if "aliases" in wb.sheetnames else []
  links_raw = _sheet_to_dicts(wb["links"]) if "links" in wb.sheetnames else []

  topics = []
  active_topic_ids = set()
  for t in topics_raw:
    if str(t.get("status", "")).strip().lower() != "active":
      continue
    topic_id = str(t.get("topic_id", "")).strip()
    active_topic_ids.add(topic_id)
    topics.append({
      "topic_id": topic_id,
      "section": (t.get("section") or "").strip(),
      "title_ru": (t.get("title_ru") or "").strip(),
      "title_en": (t.get("title_en") or "").strip(),
      "order": int(t.get("order") or 0),
      "status": "active",
      "updated_at": str(t.get("updated_at") or "").strip(),
      "notes": (t.get("notes") or "").strip(),
    })

  # tags map
  tags = []
  tag_name_by_id: dict[str, str] = {}
  for tg in tags_raw:
    tag_id = str(tg.get("tag_id", "")).strip()
    if not tag_id:
      continue
    name_ru = (tg.get("name_ru") or "").strip()
    name_en = (tg.get("name_en") or "").strip()
    tag_name_by_id[tag_id] = " ".join([tag_id, name_ru, name_en]).strip()
    tags.append({"tag_id": tag_id, "name_ru": name_ru, "name_en": name_en, "notes": (tg.get("notes") or "").strip()})

  # aliases map
  aliases = []
  aliases_by_card: dict[str, list[str]] = {}
  for a in aliases_raw:
    alias = str(a.get("alias", "")).strip()
    card_id = str(a.get("card_id", "")).strip()
    if not alias or not card_id:
      continue
    aliases_by_card.setdefault(card_id, []).append(alias)
    try:
      w = float(a.get("weight") or 0)
    except Exception:
      w = 0.0
    aliases.append({"alias": alias, "card_id": card_id, "weight": w, "notes": (a.get("notes") or "").strip()})

  # card_tags
  card_tags = []
  tags_by_card: dict[str, list[str]] = {}
  for ct in card_tags_raw:
    card_id = str(ct.get("card_id", "")).strip()
    tag_id = str(ct.get("tag_id", "")).strip()
    if not card_id or not tag_id:
      continue
    tags_by_card.setdefault(card_id, []).append(tag_id)
    card_tags.append({"card_id": card_id, "tag_id": tag_id, "notes": (ct.get("notes") or "").strip()})

  cards = []
  for c in cards_raw:
    if str(c.get("status", "")).strip().lower() != "active":
      continue
    topic_id = str(c.get("topic_id", "")).strip()
    if topic_id not in active_topic_ids:
      # краевой случай: карточка ссылается на неактивную/несущ. тему — пропускаем
      continue

    card_id = str(c.get("card_id", "")).strip()
    tag_ids = tags_by_card.get(card_id, [])
    tags_text = " ".join(tag_name_by_id.get(t, t) for t in tag_ids).strip()
    aliases_text = " ".join(aliases_by_card.get(card_id, [])).strip()

    cards.append({
      "card_id": card_id,
      "topic_id": topic_id,
      "concept": (c.get("concept") or "").strip(),
      "brief": (c.get("brief") or "").strip(),
      "example": (c.get("example") or "").strip(),
      "when_use": (c.get("when_use") or "").strip(),
      "pitfalls": (c.get("pitfalls") or "").strip(),
      "keywords": (c.get("keywords") or "").strip(),
      "status": "active",
      "updated_at": str(c.get("updated_at") or "").strip(),
      "search_text": (c.get("search_text") or "").strip(),
      "tags_text": tags_text,
      "aliases_text": aliases_text,
    })

  links = []
  for l in links_raw:
    card_id = str(l.get("card_id", "")).strip()
    if not card_id:
      continue
    links.append({
      "card_id": card_id,
      "title": (l.get("title") or "").strip(),
      "url": (l.get("url") or "").strip(),
      "kind": (l.get("kind") or "").strip(),
      "notes": (l.get("notes") or "").strip(),
    })

  return {
    "topics": topics,
    "cards": cards,
    "tags": tags,
    "card_tags": card_tags,
    "aliases": aliases,
    "links": links,
  }
